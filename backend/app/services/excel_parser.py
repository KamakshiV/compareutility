"""Excel ingestion helpers (Polars + openpyxl / xlrd)."""

from __future__ import annotations

from pathlib import Path

import polars as pl
from openpyxl import load_workbook


def read_excel_column_names(path: Path) -> list[str]:
    """
    Header row only from the first sheet — much faster than loading the full workbook
    for column validation, ingestion checks, and GET /files/.../columns.
    """
    suf = path.suffix.lower()
    if suf == ".xls":
        return [str(c) for c in read_excel_dataframe(path).columns]
    if suf not in (".xlsx", ".xlsm"):
        return [str(c) for c in read_excel_dataframe(path).columns]

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            return []
        out: list[str] = []
        for i, cell in enumerate(first):
            if cell is None or (isinstance(cell, str) and not str(cell).strip()):
                out.append(f"column_{i}")
            else:
                out.append(str(cell).strip())
        return out
    finally:
        wb.close()


def read_excel_dataframe(path: Path) -> pl.DataFrame:
    if path.suffix.lower() == ".xls":
        return pl.read_excel(path, engine="xlrd")
    return pl.read_excel(path, engine="openpyxl")
